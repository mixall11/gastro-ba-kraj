#!/usr/bin/env python3
"""Export all gastro kraj CSVs to a single Excel file with hyperlinks."""
import csv, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path('/home/michal/gastro-ba-kraj')
OUT = BASE / 'gastro-mapa-export.xlsx'

KRAJ_ORDER = ['BA','TT','NR','TN','BB','ZA','KE','PO']
MAP_URL = 'https://mixall11.github.io/gastro-ba-kraj/'

CAT_COLORS = {
    'Catering dodavatel':             'FFF3E0',  # oranžová
    'Vyvaren / catering kuchyna':     'E8F5E9',  # zelená
    'Tackaren / zavodna jedalen':     'E3F2FD',  # modrá
    'Velka restauracia / banketova sala': 'FCE4EC',  # ružová
    'Restauracna siet (3+ prevadzok)': 'F3E5F5',  # fialová
    'Statne firmy a urady':           'E0E0E0',  # sivá
}

HEADERS = [
    'Kraj', 'Kategoria', 'Firma', 'ICO', 'Adresa', 'Mesto', 'Okres',
    'Velkost / Kapacita', 'Telefon', 'Email', 'Web', 'Zona', 'Poznamka', 'Mapa'
]

COL_WIDTHS = [6, 30, 35, 12, 28, 20, 18, 38, 18, 28, 25, 8, 50, 8]

KRAJ_NAMES = {
    'BA': 'BA Bratislavsky',
    'TT': 'TT Trnavsky',
    'NR': 'NR Nitriansky',
    'TN': 'TN Trenciansky',
    'BB': 'BB Banskobystric',
    'ZA': 'ZA Zilinsky',
    'KE': 'KE Kosicky',
    'PO': 'PO Presovsky',
}

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill('solid', fgColor='1A237E')
header_font = Font(color='FFFFFF', bold=True, size=10)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

total = 0

for kraj in KRAJ_ORDER:
    csv_path = BASE / f'gastro-databaza-{kraj}-kraj.csv'
    if not csv_path.exists():
        continue
    with open(csv_path, encoding='utf-8') as f:
        rows = list(csv.DictReader(f))

    ws = wb.create_sheet(title=KRAJ_NAMES[kraj])

    # Header (bez Kraj stlpca — zbytocny ked je per-sheet)
    sheet_headers = HEADERS[1:]   # skip 'Kraj'
    sheet_widths  = COL_WIDTHS[1:]
    for col, (h, w) in enumerate(zip(sheet_headers, sheet_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(sheet_headers))}1'

    for row_num, r in enumerate(rows, 2):
        kat   = r.get('Kategoria', '')
        ico   = r.get('ICO', '').strip()
        web   = r.get('Web', '').strip()
        email = r.get('Email', '').strip()

        row_fill = PatternFill('solid', fgColor=CAT_COLORS.get(kat, 'FFFFFF'))

        cells_data = [
            ('kat',     kat),
            ('firma',   r.get('Firma', '')),
            ('ico',     ico),
            ('adresa',  r.get('Adresa', '')),
            ('mesto',   r.get('Mesto_Stvrt', '')),
            ('okres',   r.get('Okres', '')),
            ('velkost', r.get('Velkost_Trzby_Kapacita', '')),
            ('tel',     r.get('Telefon', '')),
            ('email',   email),
            ('web',     web),
            ('zona',    r.get('Zona_trasa', '')),
            ('pozn',    r.get('Poznamka', '')),
            ('mapa',    'MAPA'),
        ]

        for col, (key, val) in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = row_fill
            cell.font = Font(size=9)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=(col in (1, 7, 12)))

            if key == 'ico' and re.match(r'^\d{8}$', ico):
                cell.hyperlink = f'https://www.finstat.sk/{ico}'
                cell.font = Font(size=9, color='1565C0', underline='single')
            elif key == 'email' and '@' in email:
                cell.hyperlink = f'mailto:{email}'
                cell.font = Font(size=9, color='1565C0', underline='single')
            elif key == 'web' and web and web not in ('', 'doplnit'):
                url = web if web.startswith('http') else f'https://{web}'
                cell.hyperlink = url
                cell.font = Font(size=9, color='1565C0', underline='single')
            elif key == 'mapa':
                cell.hyperlink = MAP_URL
                cell.font = Font(size=9, color='1565C0', underline='single', bold=True)

        total += 1

    # Summary row
    ws.cell(row=len(rows) + 2, column=1, value=f'Spolu: {len(rows)} firiem')
    ws.cell(row=len(rows) + 2, column=1).font = Font(bold=True, size=9)

# Counts per category (extra sheet)
ws2 = wb.create_sheet('Prehlad')
ws2.column_dimensions['A'].width = 38
for i, k in enumerate(KRAJ_ORDER + ['SPOLU'], 2):
    ws2.column_dimensions[get_column_letter(i)].width = 10

ws2.append(['Kategoria / Kraj'] + KRAJ_ORDER + ['SPOLU'])
ws2['A1'].font = Font(bold=True)

from collections import defaultdict
counts = defaultdict(lambda: defaultdict(int))
cat_totals = defaultdict(int)
kraj_totals = defaultdict(int)

for kraj in KRAJ_ORDER:
    csv_path = BASE / f'gastro-databaza-{kraj}-kraj.csv'
    if not csv_path.exists():
        continue
    with open(csv_path, encoding='utf-8') as f:
        for r in csv.DictReader(f):
            kat = r['Kategoria']
            counts[kat][kraj] += 1
            cat_totals[kat] += 1
            kraj_totals[kraj] += 1

cat_order = [
    'Catering dodavatel', 'Vyvaren / catering kuchyna',
    'Tackaren / zavodna jedalen', 'Velka restauracia / banketova sala',
    'Restauracna siet (3+ prevadzok)', 'Statne firmy a urady'
]
for kat in cat_order:
    row = [kat] + [counts[kat].get(k, 0) for k in KRAJ_ORDER] + [cat_totals[kat]]
    ws2.append(row)
    ws2.cell(ws2.max_row, 1).fill = PatternFill('solid', fgColor=CAT_COLORS.get(kat, 'FFFFFF'))

# Totals row
ws2.append(['SPOLU'] + [kraj_totals[k] for k in KRAJ_ORDER] + [total])
ws2.cell(ws2.max_row, 1).font = Font(bold=True)

wb.save(OUT)
print(f'OK: {OUT}')
print(f'Spolu: {total} zaznamov')
